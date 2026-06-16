from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from app.models.testcase import TestCaseRequest
from app.agents.testcase_agent import testcase_agent
from app.generators.report_generator import report_generator
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io
import os

router = APIRouter(prefix="/api/v1/testcase", tags=["用例生成"])


@router.post("/generate")
async def generate_testcases(req: TestCaseRequest):
    """生成测试用例"""
    try:
        result = testcase_agent.generate(
            requirement_result=req.requirement_result,
            base_url=req.base_url,
            page_text=req.page_text,
            feedback=req.feedback,
            previous_result=req.previous_result,
        )
        return {"success": True, "data": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/save")
async def save_testcases(data: dict):
    """保存测试用例到本地"""
    try:
        filepath = report_generator.save_json(data, "testcases", "testcases")
        return {"success": True, "filepath": filepath}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/export-excel")
async def export_testcases_excel(payload: dict):
    """将测试用例导出为 Excel 文件"""
    try:
        testcases = payload.get("testcases", [])
        filename = payload.get("filename", "testcases")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # 样式定义
        header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        case_font = Font(name="微软雅黑", bold=True, size=11, color="00d4ff")
        case_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
        
        normal_font = Font(name="微软雅黑", size=10)
        normal_alignment = Alignment(vertical="top", wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="top")
        
        thin_border = Border(
            left=Side(style="thin", color="2a2a4a"),
            right=Side(style="thin", color="2a2a4a"),
            top=Side(style="thin", color="2a2a4a"),
            bottom=Side(style="thin", color="2a2a4a"),
        )

        # 列宽
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 55

        # 表头
        headers = ["序号", "用例编号", "测试步骤", "预期结果"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[1].height = 28

        # 写入数据
        row = 2
        seq = 1
        for tc in testcases:
            case_id = tc.get("case_id", "")
            title = tc.get("title", "")
            steps = tc.get("steps", [])

            if steps:
                # 有步骤：逐行写入
                for step in steps:
                    action = step.get("action", "")
                    expected = step.get("expected", "")

                    ws.cell(row=row, column=1, value=seq).font = normal_font
                    ws.cell(row=row, column=1).alignment = center_alignment
                    ws.cell(row=row, column=1).border = thin_border

                    ws.cell(row=row, column=2, value=case_id).font = Font(name="微软雅黑", bold=True, size=10)
                    ws.cell(row=row, column=2).alignment = center_alignment
                    ws.cell(row=row, column=2).border = thin_border

                    ws.cell(row=row, column=3, value=action).font = normal_font
                    ws.cell(row=row, column=3).alignment = normal_alignment
                    ws.cell(row=row, column=3).border = thin_border

                    ws.cell(row=row, column=4, value=expected).font = normal_font
                    ws.cell(row=row, column=4).alignment = normal_alignment
                    ws.cell(row=row, column=4).border = thin_border

                    row += 1
                    seq += 1
            else:
                # 无步骤：合并为一个概要行
                desc = tc.get("expected", "") or title
                ws.cell(row=row, column=1, value=seq).font = normal_font
                ws.cell(row=row, column=1).alignment = center_alignment
                ws.cell(row=row, column=1).border = thin_border

                ws.cell(row=row, column=2, value=case_id).font = Font(name="微软雅黑", bold=True, size=10)
                ws.cell(row=row, column=2).alignment = center_alignment
                ws.cell(row=row, column=2).border = thin_border

                ws.cell(row=row, column=3, value=title).font = normal_font
                ws.cell(row=row, column=3).alignment = normal_alignment
                ws.cell(row=row, column=3).border = thin_border

                ws.cell(row=row, column=4, value=desc).font = normal_font
                ws.cell(row=row, column=4).alignment = normal_alignment
                ws.cell(row=row, column=4).border = thin_border

                row += 1
                seq += 1

        # 冻结首行
        ws.freeze_panes = "A2"

        # 输出到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        safe_filename = f"{filename}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
